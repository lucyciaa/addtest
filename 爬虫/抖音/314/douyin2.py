from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook



def crawler_3():

    filename = '../点赞收藏评论数.txt'
    file = open(filename,'w',encoding='utf-8')

    options = Options()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
    browser = webdriver.Chrome(options=options)

    douyin = load_workbook(r'FY x 明洞抖音（即墨店）.xlsx')
    ws = douyin[douyin.sheetnames[0]]

    rows = ws.max_row
    prj = ws.columns
    prjTuple = tuple(prj)
    url_list = []

    for cell in prjTuple[5]:
        curr_str = str(cell.value)
        if curr_str.__contains__('http'):
            url = curr_str.split('http')[1]
            url = 'http' + url
            url_list.append(url)
    for url in url_list:
        print(url)
        browser.get(url)
        time.sleep(0.5)
        # 点赞，收藏，评论
        need = browser.find_elements(by=By.CLASS_NAME, value='CE7XkkTw')
        # 转发
        relay = browser.find_elements(by=By.CLASS_NAME, value='Uehud9DZ')
        len_need = len(need)
        util_dict = {0: '点赞', 1: '评论', 2: '收藏', 3:'转发'}
        need_dict = {'点赞': 0, '评论': 0, '收藏': 0, '转发': 0}
        for j in range(len_need):
            time.sleep(0.2)
            need_dict[util_dict[j]] = int(need[j].text)
        if(len(relay) > 0):
            for j in range(len(relay)):
                if(relay[j].text != '分享'):
                    time.sleep(0.2)
                    # print(int(relay[j].text))
                    need_dict['转发'] = int(relay[j].text)
        file.write(url)
        for key in need_dict:
            file.write(str(key) + ":" + str(need_dict[key]) + '\t')
            print(str(key) + ":" + str(need_dict[key]))

        file.write('\n')


if __name__ == '__main__':
    crawler_3()