from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook

#
# def crawler_2():
#     options = Options()
#     options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
#     browser = webdriver.Chrome(options=options)
#
#     # 输出：点赞，收藏，评论
#     need = browser.find_elements(by=By.CLASS_NAME, value='ccH7ZaBs')
#     len_need = len(need)
#     util_dict = {0:'点赞',1:'评论',2:'收藏'}
#     need_dict = {'点赞': 0, '评论': 0,'收藏': 0}
#     for j in range(len_need-1):
#         time.sleep(0.2)
#         need_dict[util_dict[j]] = int(need[j].text)
#
#     for key in need_dict:
#         print(key + ":" + str(need_dict[key]))
#
#
#     # d_close = browser.find_element(by=By.ID,value='verify-bar-close')
#     # browser.execute_script('arguments[0].click();', d_close)
#
# def crawler_1():
#     options = Options()
#     options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
#     browser = webdriver.Chrome(options=options)
#
#     file = open('url.txt', encoding='utf-8')
#     urls = file.readlines()
#     for url in urls:
#         print(url)
#         browser.get(url)
#         time.sleep(0.5)
#         # 输出：点赞，收藏，评论
#         need = browser.find_elements(by=By.CLASS_NAME, value='ccH7ZaBs')
#         len_need = len(need)
#         util_dict = {0: '点赞', 1: '评论', 2: '收藏'}
#         need_dict = {'点赞': 0, '评论': 0, '收藏': 0}
#         for j in range(len_need - 1):
#             time.sleep(0.2)
#             need_dict[util_dict[j]] = int(need[j].text)
#
#         for key in need_dict:
#             print(key + ":" + str(need_dict[key]))


def crawler_3():

    filename = './点赞收藏评论数.txt'
    file = open(filename,'w',encoding='utf-8')

    options = Options()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
    browser = webdriver.Chrome(options=options)

    douyin = load_workbook(r'FY x 魔局小红书22年(1)(1).xlsx')
    ws = douyin[douyin.sheetnames[0]]

    rows = ws.max_row
    prj = ws.columns
    prjTuple = tuple(prj)
    url_list = []
    for cell in prjTuple[3]:
        curr_str = str(cell.value)
        if curr_str.__contains__('http'):
            url = curr_str.split('http')[1].split('复制')[0]
            url = 'http' + url
            url_list.append(url)

    for url in url_list:
        print(url)
        browser.get(url)
        time.sleep(0.5)
        # 输出：点赞，收藏，评论
        need = browser.find_elements(by=By.CLASS_NAME, value='ccH7ZaBs')
        print(need[0])
        len_need = len(need)
        util_dict = {0: '点赞', 1: '评论', 2: '收藏'}
        need_dict = {'点赞': 0, '评论': 0, '收藏': 0}
        for j in range(len_need - 1):
            time.sleep(0.2)
            need_dict[util_dict[j]] = int(need[j].text)

        file.write(url)
        for key in need_dict:
            file.write(key + ":" + str(need_dict[key]) + '\t')
            print(key + ":" + str(need_dict[key]))

        file.write('\n')


if __name__ == '__main__':
    crawler_3()